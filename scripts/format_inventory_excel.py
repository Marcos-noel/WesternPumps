from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def find_col(header: list[str], aliases: list[str]) -> int | None:
    for alias in aliases:
        try:
            return header.index(alias)
        except ValueError:
            continue
    return None


def coerce_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if text == "":
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def is_category_row(desc: str, unit: Any, store_a: Any, store_b: Any, totals: Any) -> bool:
    if not desc:
        return False
    if unit or store_a or store_b or totals:
        return False
    clean = desc.strip()
    return clean.isupper() or clean.endswith(":")


def format_inventory_excel(input_path: Path, output_path: Path) -> tuple[int, int]:
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    header_row_idx = None
    header: list[str] | None = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), start=1):
        normalized = [normalize_header(c) for c in row]
        has_name_col = any(c in {"item_description", "description", "item", "name", "item_name"} for c in normalized)
        if has_name_col:
            header_row_idx = i
            header = normalized
            break

    if header_row_idx is None or header is None:
        raise ValueError("Header row not found (expected ITEM DESCRIPTION/name column).")

    idx_desc = find_col(header, ["item_description", "description", "item", "name", "item_name"])
    idx_unit = find_col(header, ["unit", "uom", "unit_of_measure"])
    idx_store_a = find_col(header, ["store_a", "storea"])
    idx_store_b = find_col(header, ["store_b", "storeb"])
    idx_totals = find_col(header, ["totals", "total", "qty", "quantity", "quantity_on_hand"])
    idx_pic = find_col(header, ["picture_links", "picture_link", "image", "image_url"])

    if idx_desc is None:
        raise ValueError("Missing item description column.")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "inventory_import"
    out_headers = [
        "sku",
        "name",
        "image_url",
        "description",
        "unit_price",
        "quantity_on_hand",
        "min_quantity",
        "tracking_type",
        "unit_of_measure",
        "category",
        "location",
        "supplier",
    ]
    out_ws.append(out_headers)

    current_category: str | None = None
    written = 0
    skipped = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        desc = str(row[idx_desc] or "").strip()
        if not desc:
            continue
        unit = row[idx_unit] if idx_unit is not None else None
        store_a = row[idx_store_a] if idx_store_a is not None else None
        store_b = row[idx_store_b] if idx_store_b is not None else None
        totals = row[idx_totals] if idx_totals is not None else None

        if is_category_row(desc, unit, store_a, store_b, totals):
            current_category = desc.strip()
            continue

        image_url = str(row[idx_pic]).strip() if (idx_pic is not None and row[idx_pic] is not None) else ""
        if not image_url:
            skipped += 1
            continue

        a = coerce_int(store_a) or 0
        b = coerce_int(store_b) or 0
        qty = coerce_int(totals)
        if qty is None:
            qty = a + b

        unit_text = str(unit).strip() if unit is not None else ""

        # Leave SKU blank so backend generates unique SKU automatically on import.
        out_ws.append(
            [
                "",
                desc,
                image_url,
                "",
                "",
                qty or 0,
                0,
                "BATCH",
                unit_text,
                current_category or "",
                "",
                "",
            ]
        )
        written += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)
    return written, skipped


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert source inventory workbook into import-ready format.")
    parser.add_argument("input", type=Path, help="Path to source .xlsx")
    parser.add_argument("output", type=Path, help="Path to output .xlsx")
    args = parser.parse_args()

    written, skipped = format_inventory_excel(args.input, args.output)
    print(f"Wrote {written} rows to {args.output}")
    print(f"Skipped {skipped} rows without image_url")


if __name__ == "__main__":
    main()
