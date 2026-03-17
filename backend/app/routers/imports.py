from __future__ import annotations

import re
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.db import get_db
from app.audit import log_audit
from app.deps import get_current_user, require_roles
from app.models import Category, Location, Part, PartLocationStock, User
from app.sku import generate_system_sku

try:
    import openpyxl
except Exception:  # pragma: no cover - handled at runtime
    openpyxl = None


router = APIRouter(prefix="/api/import", tags=["import"])


class ImportSummary(BaseModel):
    created: int
    skipped: int
    failed: int
    errors: list[str] = Field(default_factory=list)


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _coerce_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if text == "":
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _is_category_row(desc: str, unit: Any, store_a: Any, store_b: Any, totals: Any) -> bool:
    if not desc:
        return False
    if unit or store_a or store_b or totals:
        return False
    clean = desc.strip()
    return clean.isupper() or clean.endswith(":")


def _find_col(header: list[str], aliases: list[str]) -> int | None:
    for alias in aliases:
        try:
            return header.index(alias)
        except ValueError:
            continue
    return None


def _get_or_create_category(db: Session, name: str) -> int | None:
    clean = name.strip()
    if not clean:
        return None
    existing = db.scalar(select(Category).where(Category.name == clean))
    if existing:
        return existing.id
    category = Category(name=clean)
    db.add(category)
    db.commit()
    db.refresh(category)
    return category.id


def _get_or_create_location(db: Session, name: str) -> int | None:
    clean = name.strip()
    if not clean:
        return None
    existing = db.scalar(select(Location).where(Location.name == clean))
    if existing:
        return existing.id
    loc = Location(name=clean)
    db.add(loc)
    db.commit()
    db.refresh(loc)
    return loc.id


@router.post(
    "/inventory-xlsx",
    response_model=ImportSummary,
    dependencies=[Depends(require_roles("store_manager", "manager", "admin"))],
)
def import_inventory_xlsx(
    file: UploadFile = File(...),
    dry_run: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ImportSummary:
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    try:
        wb = openpyxl.load_workbook(file.file, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - runtime validation
        raise HTTPException(status_code=400, detail=f"Failed to read workbook: {exc}") from exc

    ws = wb.active
    header_row_idx = None
    header = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), start=1):
        if not row:
            continue
        normalized = [_normalize_header(c) for c in row]
        has_name_col = any(c in {"item_description", "description", "item", "name", "item_name"} for c in normalized)
        if has_name_col:
            header_row_idx = i
            header = normalized
            break

    if header_row_idx is None or header is None:
        raise HTTPException(status_code=400, detail="Header row not found (expected 'ITEM DESCRIPTION').")

    idx_desc = _find_col(header, ["item_description", "description", "item", "name", "item_name"])
    idx_unit = _find_col(header, ["unit", "uom", "unit_of_measure"])
    idx_store_a = _find_col(header, ["store_a", "storea"])
    idx_store_b = _find_col(header, ["store_b", "storeb"])
    idx_totals = _find_col(header, ["totals", "total", "qty", "quantity", "quantity_on_hand"])
    idx_pic = _find_col(header, ["picture_links", "picture_link", "image", "image_url"])

    if idx_desc is None:
        raise HTTPException(status_code=400, detail="Header must include ITEM DESCRIPTION.")

    created = 0
    skipped = 0
    errors: list[str] = []
    current_category: str | None = None

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1
    ):
        desc = str(row[idx_desc] or "").strip()
        if not desc:
            continue
        unit = row[idx_unit] if idx_unit is not None else None
        store_a = row[idx_store_a] if idx_store_a is not None else None
        store_b = row[idx_store_b] if idx_store_b is not None else None
        totals = row[idx_totals] if idx_totals is not None else None
        pic = row[idx_pic] if idx_pic is not None else None

        if _is_category_row(desc, unit, store_a, store_b, totals):
            current_category = desc.strip()
            continue

        a = _coerce_int(store_a) or 0
        b = _coerce_int(store_b) or 0
        qty = _coerce_int(totals)
        if qty is None:
            qty = a + b

        sku = generate_system_sku(db)
        category_id = _get_or_create_category(db, current_category) if current_category else None
        unit_text = str(unit).strip() if unit is not None else ""
        image_url = str(pic).strip() if pic is not None else ""
        if not image_url:
            skipped += 1
            errors.append(f"Row {row_idx}: image_url is required.")
            continue

        part = Part(
            sku=sku,
            name=desc,
            description=None,
            image_url=image_url,
            unit_price=None,
            quantity_on_hand=qty or 0,
            min_quantity=0,
            tracking_type="BATCH",
            unit_of_measure=unit_text or None,
            category_id=category_id,
            location_id=None,
            supplier_id=None,
        )

        if dry_run:
            created += 1
            continue

        db.add(part)
        try:
            db.commit()
            created += 1
            if a or b:
                if a:
                    loc_a = _get_or_create_location(db, "Store A")
                    if loc_a:
                        db.add(PartLocationStock(part_id=part.id, location_id=loc_a, quantity_on_hand=a))
                if b:
                    loc_b = _get_or_create_location(db, "Store B")
                    if loc_b:
                        db.add(PartLocationStock(part_id=part.id, location_id=loc_b, quantity_on_hand=b))
                db.commit()
        except IntegrityError as exc:
            db.rollback()
            skipped += 1
            errors.append(f"Row {row_idx}: {exc}")
        except Exception as exc:  # pragma: no cover - safety net
            db.rollback()
            skipped += 1
            errors.append(f"Row {row_idx}: {exc}")

    failed = len(errors)
    log_audit(
        db,
        current_user,
        action="import",
        entity_type="inventory",
        detail={
            "file_name": file.filename,
            "dry_run": dry_run,
            "created": created,
            "skipped": skipped,
            "failed": failed,
        },
    )
    db.commit()
    return ImportSummary(created=created, skipped=skipped, failed=failed, errors=errors[:50])
